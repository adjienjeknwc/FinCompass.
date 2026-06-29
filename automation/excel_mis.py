"""
FinCompass - Excel MIS Report Automation Module
===============================================

This module automates the generation of a highly formatted, multi-tab Excel Management
Information System (MIS) report using `openpyxl`.

Tabs created:
1. "Bank Performance": Details on complaints count, resolution rate, average resolution
   times, and supervisory risk levels with conditional formatting.
2. "Category Breakdown": Metrics categorized by primary and subcategory labels.
3. "Geographic Analysis": Distribution of complaints across Indian states.

Formatting features:
- Solid fill header rows (Navy background with white text).
- Double bottom border on headers.
- Auto-fitted column widths.
- Conditional formatting: Cells under 'Risk Level' are filled with Green/Amber/Red
  according to the risk level of the bank.
- Saved to: `reports/FinCompass_MIS_Report.xlsx`
"""

import sqlite3
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path("/Users/aditi/.gemini/antigravity/scratch/FinCompass")
DB_PATH = PROJECT_ROOT / "database" / "fincompass.db"
OUTPUT_XLSX = PROJECT_ROOT / "reports" / "FinCompass_MIS_Report.xlsx"

def generate_excel_mis():
    """Generates the multi-tab formatted Excel MIS report."""
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    
    # 1. Fetch data from DB
    with sqlite3.connect(str(DB_PATH)) as conn:
        # Tab 1 data: Bank Performance
        # Join with risk scoring logic (we re-calculate or load from risk scores)
        # For simplicity, we calculate directly here
        bank_perf = pd.read_sql_query("""
            SELECT 
                b.bank_name as "Bank Name",
                b.bank_type as "Bank Type",
                COUNT(c.complaint_id) as "Total Complaints",
                SUM(CASE WHEN c.status = 'Resolved' THEN 1 ELSE 0 END) as "Resolved Count",
                ROUND((CAST(SUM(CASE WHEN c.status = 'Resolved' THEN 1 ELSE 0 END) AS REAL) / COUNT(c.complaint_id)) * 100, 1) as "Resolution Rate (%)",
                ROUND(AVG(CASE WHEN c.status = 'Resolved' THEN c.resolution_days ELSE NULL END), 1) as "Avg Resolution Days"
            FROM complaints c
            JOIN banks b ON c.bank_id = b.bank_id
            GROUP BY b.bank_name, b.bank_type
            ORDER BY "Total Complaints" DESC
        """, conn)
        
        # Add a dummy Risk Level column for conditional formatting
        # Based on average resolution days: >40 is High, >30 is Medium, else Low
        def get_risk(days):
            if days > 40: return "High"
            elif days > 29: return "Medium"
            else: return "Low"
        bank_perf["Risk Level"] = bank_perf["Avg Resolution Days"].apply(get_risk)

        # Tab 2 data: Category Breakdown
        category_breakdown = pd.read_sql_query("""
            SELECT 
                cat.category_name as "Category",
                cat.subcategory_name as "Subcategory",
                COUNT(c.complaint_id) as "Complaint Count",
                ROUND((CAST(SUM(CASE WHEN c.status = 'Resolved' THEN 1 ELSE 0 END) AS REAL) / COUNT(c.complaint_id)) * 100, 1) as "Resolution Rate (%)"
            FROM complaints c
            JOIN categories cat ON c.subcategory_id = cat.subcategory_id
            GROUP BY cat.category_name, cat.subcategory_name
            ORDER BY "Category" ASC, "Complaint Count" DESC
        """, conn)

        # Tab 3 data: Geographic Analysis
        geo_analysis = pd.read_sql_query("""
            SELECT 
                c.state as "State / UT",
                COUNT(c.complaint_id) as "Complaint Count",
                SUM(CASE WHEN c.status = 'Resolved' THEN 1 ELSE 0 END) as "Resolved Count",
                ROUND((CAST(SUM(CASE WHEN c.status = 'Resolved' THEN 1 ELSE 0 END) AS REAL) / COUNT(c.complaint_id)) * 100, 1) as "Resolution Rate (%)"
            FROM complaints c
            GROUP BY c.state
            ORDER BY "Complaint Count" DESC
        """, conn)

    # 2. Build workbook using openpyxl
    wb = openpyxl.Workbook()
    
    # Setup styles
    font_title = Font(name="Arial", size=14, bold=True, color="002855")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Arial", size=10)
    font_bold = Font(name="Arial", size=10, bold=True)
    
    fill_navy = PatternFill(start_color="002855", end_color="002855", fill_type="solid")
    fill_light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Conditional formatting fills
    fill_red = PatternFill(start_color="FFD1D1", end_color="FFD1D1", fill_type="solid")
    fill_amber = PatternFill(start_color="FFEAA7", end_color="FFEAA7", fill_type="solid")
    fill_green = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_thin = Side(border_style="thin", color="D3D3D3")
    border_double = Side(border_style="double", color="000000")
    
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    header_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_double)

    # Helper function to style a sheet
    def write_sheet(ws, title, df, is_bank_sheet=False):
        # Set Grid lines visible
        ws.views.sheetView[0].showGridLines = True
        
        # 1. Add Sheet Title block
        ws.cell(row=1, column=1, value=title).font = font_title
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 10  # blank buffer row
        
        # 2. Add Header row
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = align_center
            cell.border = header_border
        ws.row_dimensions[3].height = 25
        
        # 3. Add Data rows
        for row_num, row_data in enumerate(df.values, 4):
            ws.row_dimensions[row_num].height = 18
            for col_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=val)
                cell.font = font_data
                cell.border = cell_border
                
                # Check data types and align
                if isinstance(val, (int, float)):
                    cell.alignment = align_right
                    # Number format formatting
                    if "Rate" in headers[col_num-1]:
                        cell.number_format = '0.0"%"'
                    elif "Days" in headers[col_num-1] or "Count" in headers[col_num-1] or "Complaints" in headers[col_num-1]:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = align_left
                    
                # Apply conditional formatting to Risk Level column
                if is_bank_sheet and headers[col_num-1] == "Risk Level":
                    cell.alignment = align_center
                    if val == "High":
                        cell.fill = fill_red
                        cell.font = Font(name="Arial", size=10, bold=True, color="9C0006")
                    elif val == "Medium":
                        cell.fill = fill_amber
                        cell.font = Font(name="Arial", size=10, bold=True, color="9C6500")
                    else:
                        cell.fill = fill_green
                        cell.font = Font(name="Arial", size=10, bold=True, color="006100")
                        
        # 4. Add Summary Row (total sums or averages)
        total_row_idx = len(df) + 4
        ws.row_dimensions[total_row_idx].height = 20
        ws.cell(row=total_row_idx, column=1, value="Total / System Average").font = font_bold
        ws.cell(row=total_row_idx, column=1).border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
        
        for col_num in range(2, len(headers) + 1):
            cell = ws.cell(row=total_row_idx, column=col_num)
            cell.font = font_bold
            cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
            col_letter = get_column_letter(col_num)
            
            header_name = headers[col_num-1]
            if "Complaints" in header_name or "Count" in header_name:
                cell.value = f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})"
                cell.number_format = '#,##0'
                cell.alignment = align_right
            elif "Rate" in header_name or "Days" in header_name:
                cell.value = f"=AVERAGE({col_letter}4:{col_letter}{total_row_idx-1})"
                cell.number_format = '0.0'
                cell.alignment = align_right
                
        # 5. Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Avoid title formulas length bloating column widths
                if cell.row == 1:
                    continue
                val_str = str(cell.value or '')
                if cell.row == total_row_idx:
                    val_str = "Total / System Average" # Approximate formula text
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Tab 1: Bank Performance
    ws1 = wb.active
    ws1.title = "Bank Performance"
    write_sheet(ws1, "Supervisory Bank Performance Metrics", bank_perf, is_bank_sheet=True)
    
    # Tab 2: Category Breakdown
    ws2 = wb.create_sheet(title="Category Breakdown")
    write_sheet(ws2, "Complaint Breakdown by Category & Subcategory", category_breakdown)
    
    # Tab 3: Geographic Analysis
    ws3 = wb.create_sheet(title="Geographic Analysis")
    write_sheet(ws3, "Complaint Distribution Across Indian States", geo_analysis)
    
    # Save Workbook
    wb.save(OUTPUT_XLSX)
    print(f"Excel MIS Report successfully generated and saved to: {OUTPUT_XLSX}")


if __name__ == "__main__":
    generate_excel_mis()
